import pandas as pd
import openpyxl
import holitest as hlt
import numpy as np
import os

rnn_path_fig = '/tmp/RNN_python/figures_seq2seq_test/'
# rnn_path_output_data = '/tmp/RNN_python/output_data_test/'
rnn_path_output_data = '/tmp/RNN_python/learning_length=10num_units50output_data_test/'

sarima_path_fig = '/tmp/RNN_python/figures_sarima/'
sarima_path_output_data = '/tmp/RNN_python/output_data_sarima/'

path_output_data = '/tmp/RNN_python/evaluate_data1/'

os.makedirs(path_output_data, exist_ok=True)
eval_data_set_kari = hlt.eval_series_data()
eval_data_set = eval_data_set_kari#[:700]

rnn_predicted_traffic_data = pd.read_excel(
                                    rnn_path_output_data + 'seq2seq_predict.xlsx',
                                    index_col = [0, 1, 2],
                                    header = 0
                                    )

rnn_log_gauss_error_data = pd.read_excel(
                                    rnn_path_output_data + 'seq2seq_error_gauss.xlsx',
                                    index_col = 0,
                                    header = 0
                                    )

rnn_series_error_data = pd.read_excel(
                                    rnn_path_output_data + 'seq2seq_error_p_h.xlsx',
                                    index_col = [0, 1, 2],
                                    header = 0
                                    )

sarima_predicted_traffic_data = pd.read_excel(
                                    sarima_path_output_data + 'sarima_predict1.xlsx',
                                    index_col = [0, 1, 2],
                                    header = 0
                                    )

sarima_error_gaus_data = pd.read_excel(
                                    sarima_path_output_data + 'sarima_error_gauss.xlsx',
                                    index_col = 0,
                                    header = 0
                                    )

s_arima_err_p_h_data = pd.read_excel(
                                    sarima_path_output_data + 's_arima_err_p_h_data1.xlsx',
                                    index_col = [0, 1, 2],
                                    header = 0
                                    )


rnn_learning_data_day_len = 10
sarima_input_data_day_len = 10
thrd = 54.5
num_nom = abs((rnn_learning_data_day_len - sarima_input_data_day_len) * 24)

# define real series data length
# if rnn_learning_data_day_len > sarima_input_data_len:
#     real_day_data = eval_data_set[sarima_input_data_day_len * 24:]
#
# else:
#     real_day_data = eval_data_set.values.reshape(-1)[rnn_learning_data_day_len * 24:]

rnn_log_gauss_error = rnn_log_gauss_error_data.values
#rnn erorr detect
check = 0
error_day_data = pd.DataFrame()
error_rnn_data = []
error_sarima_data = []

for i in range(len(rnn_log_gauss_error)):
    dataframe_2_ = eval_data_set[(rnn_learning_data_day_len + i) * 24: \
        (rnn_learning_data_day_len + i) * 24 + 24]
    # print(rnn_log_gauss_error[i])
    # print(i)

    if rnn_log_gauss_error[i] < -thrd:
        # if check == 0:
        #     dataframe_2_.to_excel(path_output_data + 'rnn_anom_day.xlsx')
        #     wb = openpyxl.load_workbook(path_output_data + 'rnn_anom_day.xlsx', False)
        #     sheet = wb.active
        #     sheet.cell(row = 1, column = 5, value = 'rnn_forecast')
        #     sheet.cell(row = 1, column = 6, value = 'sarima_forecast')
        #
        #     # wb.save(path_output_data + 'rnn_anom_day.xlsx')
        #
        # else:
        #     # wb = openpyxl.load_workbook(path_output_data + 'rnn_anom_day.xlsx')
        #     # sheet = wb.active
        #
        #     for k in range(24):
        #         # sheet.cell(row = check*24 + 1 + k, column = 1, \
        #         #     value=str(eval_data_set[i*24: i*24 + 1].index.get_level_values('day_data')))
        #         # sheet.cell(row = check*24 + 1 + k, column = 2, \
        #         #     value=str(eval_data_set[i*24: i*24 + 1].index.get_level_values('day')))
        #         if k == 0:
        #             index_name = str(eval_data_set[i*24 + k: i*24 + k + 1].index.get_level_values('date-time'))[16:26]
        #         else:
        #             index_name = str(eval_data_set[i*24 + k: i*24 + k + 1].index.get_level_values('date-time'))[16:35]
        #         sheet.cell(row = check*24 + 2 + k, column = 3, \
        #             value=index_name)
        #         sheet.cell(row = check*24 + 2 + k, column = 4, \
        #             value=eval_data_set.values.reshape(-1)[i*24 + k + 1])
        #     # wb.save(path_output_data + 'rnn_anom_day.xlsx')
        #
        # # wb = openpyxl.load_workbook(path_output_data + 'rnn_anom_day.xlsx')
        # # sheet = wb.active
        error_day_data = error_day_data.append(dataframe_2_)
        # print(len(error_day_data))
        error_rnn_data.append(rnn_predicted_traffic_data.values[i*24: i*24 + 24, 0])
        error_sarima_data.append(sarima_predicted_traffic_data.values[i*24: i*24 + 24])
        # print(error_day_data)
        print(error_rnn_data)
        print(len(error_day_data), len(error_rnn_data))

error_day_data = error_day_data.assign(\
    rnn_forecast_number = np.reshape(error_rnn_data, -1),
    sarima_forecast_number = np.reshape(error_sarima_data, -1)
    )

    # print(dataframe_2_)

error_day_data.to_excel(path_output_data + 'rnn_anom_day.xlsx')
    # wb = openpyxl.load_workbook(path_output_data + 'rnn_anom_day.xlsx', False)
    # sheet = wb.active
    # sheet.cell(row=1, column=5, value='rnn_forecast_number')
    #
    #     for k in range(24):
    #         sheet.cell(row = check*24 + 1 + k, column=5, \
    #             value=rnn_predicted_traffic_data.values[:, 0][i*24 + k])
    #         # sheet.cell(row = i*24 + 1 + k, column=6, \
    #         #     value=sarima_predicted_traffic_data.values[num_nom + i*24 + k])
    #     check += 1
    #     # print(check)
# wb.save(path_output_data + 'rnn_anom_day.xlsx')
