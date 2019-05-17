import numpy as np
import tensorflow as tf
import pandas as pd
import os
import pathlib
import holitest as hlt
import s_arima_mod as sam
import tensorflow.contrib.eager as tfe
from tensorflow.contrib import rnn
from tensorflow.contrib import seq2seq
from tensorflow.contrib import cudnn_rnn
from tensorflow.nn import rnn_cell
# from tensorflow.python import debug as tf_debug

# import make_slide_win as msw
#import cufflinks as cf
import matplotlib.pyplot as plt
#import matplotlib as mpl
import scipy.stats as spy
import openpyxl
# import numba
# import ppp
# import plotting
# from numba import cuda
from scipy.stats import norm
#import pdb; pdb.set_trace()
from sklearn.utils import shuffle
from sklearn.model_selection import train_test_split
#start = time.time()

learning_rate = 0.001

learning_data_day_len = 10
input_digits = 24 * 5
output_digits = 24
n_hidden = 150
epochs = 100
batch_size = 30

ample = 0
# day = 'Tue'
# learning_length = 700
thrd = 54.5
input_len = 24 * 10

tf.reset_default_graph()

# tfe.enable_eager_execution()


path_fig = '/tmp/RNN_python/figures_seq2seq_test/'
path_output_data = '/tmp/RNN_python/output_data_test/'
LOG_DIR = '/tmp/RNN_python/rnn_log'

class TimeSeriesDataSet:
#時系列データの時間の設定
    def __init__(self, dataframe):
        self.feature_count = len(dataframe.columns)
        self.series_length = len(dataframe)
        self.series_data = dataframe.astype('float32')

    def __getitem__(self, n):
        return TimeSeriesDataSet(self.series_data[n])

    def __length_of_sequences__(self):
        return len(self.series_data)

    def times(self):
        return self.series_data.index
#データの切り出し
    def next_batch(self, input_digits, output_digits, ample):
        data = []
        target = []

        n_index = self.series_length - (input_digits + output_digits)
        noise = ample * np.random.uniform(low=-1.0, high=1.0, \
            size=[self.series_length, self.feature_count])
        value = self.series_data.values
        noise_value = value + noise

        for i in range(0, n_index):

            data.append(noise_value[i: i + input_digits])
            target.append(spy.zscore(value[i+input_digits: i+input_digits+output_digits]))

        X = np.stack(data)
        std_Y = np.stack(target)

        #import pdb; pdb.set_trace()
        return X, std_Y

    def append(self, data_point):
        dataframe = pd.DataFrame(data_point, columns=self.series_data.columns)
        self.series_data = self.series_data.append(dataframe)

    def tail(self, n):
        return TimeSeriesDataSet(self.series_data.tail(n))

    def as_array(self):
        return np.stack([self.series_data.as_matrix()])
    #標準化
    def mean(self):
        return self.series_data.mean()

    def std(self):
        return self.series_data.std()

    def standardize(self, mean = None, std = None):
        if mean is None:
            mean = self.mean()
        if std is None:
            std = self.std()
        return TimeSeriesDataSet((self.series_data - mean) / std)


class Early_Stopping():
    def __init__(self, patience=0, verbose=0):
        self._step = 0
        self._loss = float('inf')
        self.patience = patience
        self.verbose = verbose

    def validate(self, loss):
        if self._loss < loss:
            self._step += 1
            if self._step > self.patience:
                if self.verbose:
                    print('early stopping')
                return True
        else:
            self._step = 0
            self._loss = loss

        return False


def mape_evaluation(p_data, day_d):
    return np.sum(abs((p_data - day_d)/day_d), axis = 0)*100/24

def normdist(x,mu,sigma):
    # ndarray xに対する平均mu, 標準偏差sigmaの正規分布の確率密度関数を返す関数
    return np.array([norm.pdf(x = x[i], loc = mu, scale = sigma) for i in range(len(x)) ])

eval_data_set_kari = hlt.eval_series_data()
eval_data_set = eval_data_set_kari
eval_data_set_inst = TimeSeriesDataSet(eval_data_set)
eval_series_length = eval_data_set_inst.series_length
(eval_X, eval_Y) = eval_data_set_inst.next_batch(input_digits = input_digits, \
    output_digits=output_digits, ample = ample)

n_in = len(eval_X[0][0])
n_out = len(eval_Y[0][0])

N_train = int((learning_data_day_len * 24 - (input_digits + output_digits))* 0.95)
N_validation = (learning_data_day_len * 24- (input_digits + output_digits)) - N_train
n_batches = N_train // batch_size

num_day = eval_series_length // 24 - 1

dataframe_2_  = []
day_d = []
series_error = []
gauss_error = []
log_gauss_error = []
rnn_day_series_mape = []
p_data_sr = []
rnn_np_p_data_sr = np.empty(0, int)
an_d = 1
check = 0
num_of_err = 0

anom_day_fig = plt.figure(figsize=(15, 25))

for k in range(0, (eval_series_length - (learning_data_day_len * 24 + output_digits)) // 24 - 1):
    tf.reset_default_graph()
    def inference(x, y, n_batch, is_training,
                  input_digits=None,
                  output_digits=None,
                  n_hidden=None,
                  n_out=None):
        def weight_variable(shape):
            initial = tf.truncated_normal(shape, stddev=0.01)
            return tf.Variable(initial)

        def bias_variable(shape):
            initial = tf.zeros(shape, dtype=tf.float32)
            return tf.Variable(initial)

        def batch_normalization(shape, x):
            with tf.name_scope('batch_normalization'):
                eps = 1e-8
                # beta = tf.Variable(tf.zeros(shape))
                # gamma = tf.Variable(tf.ones(shape))
                mean, var = tf.nn.moments(x, [0, 1])
                # nom_batch = gamma * (x - mean) / tf.sqrt(var + eps) + beta
                nom_batch = (x - mean) / tf.sqrt(var + eps)
                # print(nom_batch[0], len(nom_batch[0]))
                return nom_batch

        encoder = rnn_cell.GRUCell(n_hidden)
        encoder_outputs = []
        encoder_states = []

        # Encode
        # encoder = cudnn_rnn.CudnnGRU(
        #                             num_layers=1,
        #                             num_units=int(n_hidden),
        #                             input_mode='auto_select',
        #                             # direction='bidirectional',
        #                             dtype=tf.float32)

        state = encoder.zero_state(n_batch, tf.float32)

        # [input_digits, n_batch, 1], [1, n_batch, n_hidden]
        # encoder_outputs, encoder_states = \
        #     encoder(tf.reshape(batch_normalization(input_digits, x), \
        #                 [input_digits, n_batch, n_in]),
        #             # initial_state = state,
        #             training = True
        #             )

        with tf.variable_scope('Encoder'):
            for t in range(input_digits):
                if t > 0:
                    tf.get_variable_scope().reuse_variables()
                (output, state) = encoder(batch_normalization(input_digits, x)[:, t, :], state)
                encoder_outputs.append(output)
                encoder_states.append(state)


        # encoder = seq2seq.AttentionWrapper(encoder,
        #                                     attention_mechanism = AttentionMechanism,
        #                                     attention_layer_size = 128,
        #                                     initial_cell_state = \
        #                                     AttentionWrapper.zero_state(n_batch, tf.float32))

        # Decode


        AttentionMechanism = seq2seq.BahdanauAttention(num_units=100,
                                                        memory=tf.reshape(encoder_outputs, \
                                                            [n_batch, input_digits, n_hidden * 1])
                                                        )
                                                        # when use bidirectional, n_hidden * 2
                                                        # tf.reshape(encoder_outputs, n_batch, input_digits, ),
                                                        # memory_sequence_length = input_digits)
                                                        # normalize=True)


        decoder = rnn_cell.GRUCell(n_hidden)
        decoder = seq2seq.AttentionWrapper(decoder,
                                           attention_mechanism = AttentionMechanism,
                                           attention_layer_size = 50,
                                           output_attention = False)
                                           # initial_cell_state = encoder_states[-1])こいつが悪い


        state = decoder.zero_state(n_batch, tf.float32)\
            .clone(cell_state=tf.reshape(encoder_states[-1], [n_batch, n_hidden]))
        # state = encoder_states[-1]
        # decoder_outputs = tf.reshape(encoder_outputs[-1,　:,　:], [n_batch, 1])
        # [input_len, n_batch, n_hidden]
        # なんでかスライスだけエラーなし？
        decoder_outputs = [encoder_outputs[-1]]
        # decoder_outputs = [encoder_outputs[-1]]
        # 出力層の重みとバイアスを事前に定義
        V = weight_variable([n_hidden, n_out])
        c = bias_variable([n_out])
        outputs = []

        # decoder = seq2seq.BasicDecoder(cell = decoder,
        #                                 heiper = helper,
        #                                 initial_state=state,
        #                                 )


        with tf.variable_scope('Decoder'):
            for t in range(1, output_digits):
                if t > 1:
                    tf.get_variable_scope().reuse_variables()

                if is_training is True:
                    (output, state) = decoder(batch_normalization(output_digits, y)[:, t-1, :], state)
                else:
                    # 直前の出力を求める
                    out = tf.matmul(decoder_outputs[-1], V) + c
                    # elems = decoder_outputs[-1], V , c
                    # out = tf.map_fn(lambda x: x[0] * x[1] + x[2], elems)
                    # out = decoder_outputs
                    outputs.append(out)
                    (output, state) = decoder(out, state)

                # decoder_outputs.append(output)
                decoder_outputs = tf.concat([decoder_outputs, tf.reshape(output, [1, n_batch, n_hidden])], axis = 0)
                # decoder_outputs = tf.concat([decoder_outputs, output], 1)
        if is_training is True:
            output = tf.reshape(tf.concat(decoder_outputs, axis=1),
                                [-1, output_digits, n_hidden])
            with tf.name_scope('check'):
                linear = tf.einsum('ijk,kl->ijl', output, V, ) + c
                return linear
        else:
            # 最後の出力を求める
            linear = tf.matmul(decoder_outputs[-1], V) + c
            outputs.append(linear)

            output = tf.reshape(tf.concat(outputs, axis=1),
                                [-1, output_digits, n_out])
            return output

    def loss(y, t):
        with tf.name_scope('loss'):
            mse = tf.reduce_mean(tf.square(y - t), axis = [1, 0])
            # mse = tf.reduce_mean(tf.square(y - t), [1, 0])
            return mse

    def training(loss, learning_rate):
        with tf.name_scope('train_step'):
            optimizer = \
                tf.train.AdamOptimizer(learning_rate=learning_rate, beta1=0.9, beta2=0.999)

            train_step = optimizer.minimize(loss)
            return train_step


    x = tf.placeholder(tf.float32, shape=[None, input_digits, n_in])
    t = tf.placeholder(tf.float32, shape=[None, output_digits, n_out])
    n_batch = tf.placeholder(tf.int32, shape=[])
    is_training = tf.placeholder(tf.bool)

    y = inference(x, t, n_batch, is_training,
                  input_digits=input_digits,
                  output_digits=output_digits,
                  n_hidden=n_hidden, n_out=n_out)
    loss = loss(y, t)
    train_step = training(loss = loss, learning_rate = learning_rate)

    with tf.name_scope('initial'):
        init = tf.global_variables_initializer()
        sess = tf.Session()
        # sess = tf_debug.LocalCLIDebugWrapperSession(sess)
        # sess = tf_debug.TensorBoardDebugWrapperSession(sess, 'localhost:6064')
        if k == 0:
            tf.summary.FileWriter(LOG_DIR, sess.graph)
        sess.run(init)

    history = {
        'val_loss': [],
        'val_acc': []
    }

    input_data = eval_X[k * 24: (k + learning_data_day_len - 1) * 24]
    true_data = eval_Y[k * 24: (k + learning_data_day_len - 1) * 24]

    # print(input_data[:3])
    input_data_train, input_data_validation, true_data_train, \
        true_data_validation = train_test_split(input_data, true_data, \
            test_size = N_validation)

    # print('input_data_train = ', input_data_train[1:4])
    # print(len(input_data_train[0]))
    for epoch in range(epochs):
        X_, Y_ = shuffle(input_data_train, true_data_train)

        with tf.name_scope('train'):
            for h in range(n_batches):
                start = h * batch_size
                end = start + batch_size
                # print('begin learning')
                # print(y)
                # t = Y_[start:end]
                # y = inference(X_[start:end], Y_[start:end], batch_size, True,
                #               input_digits=input_digits,
                #               output_digits=output_digits,
                #               n_hidden=n_hidden, n_out=n_out)
                # loss = loss(y, t)
                #
                # train_step(loss, learning_rate)
                sess.run(train_step, feed_dict={
                    x: X_[start:end],
                    t: Y_[start:end],
                    n_batch: batch_size,
                    is_training: True
                })

        val_loss = loss.eval(session=sess, feed_dict={
            x: input_data_validation,
            t: true_data_validation,
            n_batch: N_validation,
            is_training: False
        })

        #mean_val, var_val = tf.nn.moments(X_validation, [0, 1])
        #std_val_loss = val_loss * tf.sqrt(var_val) + mean_val
        #std_val_loss = val_loss / train_std**2

        history['val_loss'].append(val_loss)
        print('epoch:', epoch,
              ' validation loss:', val_loss)

    #forcasting
    predicted_traffic = [[None] * len(eval_data_set.columns) \
    for l in range(input_digits)]

    fc_input = eval_X[learning_data_day_len * 24 - (input_digits - k * 24)].reshape(1, input_digits, 1)
    std_fc_input = spy.zscore(fc_input, axis = 1)

    z_ = std_fc_input.reshape(1, input_digits, 1)

    std_output = y.eval(session=sess, feed_dict={
        x: z_,
        n_batch: 1,
        is_training: False
    })

    tf.reset_default_graph()

    fc_input_mean = fc_input.mean(axis=1, keepdims=True)
    fc_input_std = fc_input.std(axis=1, keepdims=True)

    fc_output = std_output * fc_input_std + fc_input_mean
    fc_seq = fc_output.reshape(-2)
    print(fc_seq)
    rnn_np_p_data_sr = np.append(rnn_np_p_data_sr, fc_seq.reshape(-1), axis = 0)

    dataframe_2_ = eval_data_set[(learning_data_day_len + k) * 24: \
        (learning_data_day_len + k) * 24 + 24]
    day_d = dataframe_2_.values.reshape(-1)
    print(day_d)

    if len(day_d) != 24:
        break

    series_error.append(fc_seq - day_d)
    print(series_error)
    gauss_error.append(np.prod(normdist(np.reshape(series_error[k], -1),0,50)))
    print(gauss_error)
    log_gauss_error.append(np.log10(gauss_error[k]))

    day_mape = mape_evaluation(fc_seq, day_d)
    rnn_day_series_mape.append(day_mape)
    print(log_gauss_error[k])

    if log_gauss_error[k] < -thrd and an_d < 16:
        day_predicted_traffic = pd.DataFrame(fc_seq, \
            columns = dataframe_2_.columns, index = dataframe_2_.index)
        day_real_traffic = pd.DataFrame(day_d,\
            columns = dataframe_2_.columns, index = dataframe_2_.index)
        with pd.ExcelWriter(path_output_data + 'anom_data' + str(an_d) + '.xlsx') as writer:
            day_predicted_traffic.to_excel(writer, sheet_name = 'predict')
            day_real_traffic.to_excel(writer, sheet_name = 'real')
        print('likelihood_anormal :' , dataframe_2_[0: 1])
        fig_name = str(dataframe_2_[0:1].index.values[0])
        name = fig_name[7: 13] + fig_name[25:39]

        day_ax = anom_day_fig.add_subplot(7, 3, an_d) #sharex=True, sharey=True)
        # sarima = day_ax.plot(sarima_p_data_sr[learning_data_day_len - input_len // 24 + k]\
            # , 'r', label = 'SARIMA')
        rnn = day_ax.plot(fc_seq, label = 'RNN')
        real = day_ax.plot(day_d,label = 'real', linestyle = 'dotted')
        day_ax.legend()
        day_ax.set_title(name)
        an_d += 1

# plt.savefig(path_fig + 'anormal_day_' + str(an_d) + '_' + str(input_digits) +'.eps')
anom_day_fig.savefig(path_fig + 'rnn_seq2seq_lkhd_anormal_day.eps', dpi=250)

# predicted_traffic_data = pd.DataFrame(rnn_np_p_data_sr, \
#     columns = eval_data_set[learning_data_day_len * 24:learning_data_day_len * 24 + len(rnn_np_p_data_sr)].columns, \
#     index=eval_data_set[learning_data_day_len * 24:learning_data_day_len * 24 + len(rnn_np_p_data_sr)].index)

predicted_traffic_data = pd.DataFrame(rnn_np_p_data_sr, \
    columns = eval_data_set[learning_data_day_len * 24:learning_data_day_len * 24 + len(rnn_np_p_data_sr)].columns, \
    index=eval_data_set[learning_data_day_len * 24:learning_data_day_len * 24 + len(rnn_np_p_data_sr)].index)

predicted_traffic_data.to_excel(path_output_data + 'seq2seq_predict.xlsx')

log_gauss_error_data = pd.DataFrame(log_gauss_error)

log_gauss_error_data.to_excel(path_output_data + 'seq2seq_error_gauss.xlsx')

series_error_data = pd.DataFrame(np.reshape(series_error, -1), \
    columns = eval_data_set[learning_data_day_len * 24:learning_data_day_len * 24 + len(rnn_np_p_data_sr)].columns, \
    index=eval_data_set[learning_data_day_len * 24:learning_data_day_len * 24 + len(rnn_np_p_data_sr)].index)

series_error_data.to_excel(path_output_data + 'seq2seq_error_p_h.xlsx')

wb = openpyxl.load_workbook(path_output_data + '/seq2seq_predict.xlsx')
sheet = wb['Sheet1']
sheet.cell(row=1, column=5, value='real_number')
real_day_data = eval_data_set.values.reshape(-1)[learning_data_day_len * 24:]
for i in range(len(real_day_data)):
    sheet.cell(row = 2 + i, column=5, value=real_day_data[i])
wb.save(path_output_data + '/seq2seq_predict.xlsx')

print()
print(fc_seq)
print(len(log_gauss_error))
print()
print(rnn_day_series_mape)
print(len(rnn_day_series_mape))
print()

fig6, ax6 = plt.subplots()
ax6.plot(range(len(rnn_day_series_mape)), \
    np.array(rnn_day_series_mape, dtype = float).reshape(-1), label='mape')
ax6_1 = ax6.twinx()  # 2つのプロットを関連付ける
ax6_1.plot(range(len(log_gauss_error)), log_gauss_error, 'r', label='likelihood')
plt.savefig(path_fig + 'mape_lkhd' + str(input_digits) + '.eps', dpi=300)

# likelihood
fig3 = plt.figure(figsize=(20, 8))
ax3 = fig3.add_subplot(1,1,1)
rnn = ax3.plot(range(len(log_gauss_error)), log_gauss_error, label='rnn_likelihood')
# sarima = ax3.plot(range(len(s_arima_log_lkhd)), s_arima_log_lkhd , label = 'sarima_likelihood')
ax3.legend()
ax3.hlines([-thrd], 0, len(log_gauss_error), "r", linestyles='dashed')
ax3.set_xlabel('day')
ax3.set_ylabel('likelihood')
fig3.savefig(path_fig + 'seq2seq_sarima_lkhd.eps', dpi=280)

# mape
fig7 = plt.figure(figsize=(20, 8))
ax7 = fig7.add_subplot(1,1,1)
rnn = ax7.plot(range(len(rnn_day_series_mape)), np.array(rnn_day_series_mape, \
    dtype = float).reshape(-1), label='rnn_mape')
# sarima = ax7.plot(range(len(s_arima_day_series_mape)), s_arima_day_series_mape, label = 'sarima_mape')
ax7.legend()
ax7.set_xlabel('day')
ax7.set_ylabel('likelihood')
fig7.savefig(path_fig + 'seq2seq_sarima_mape.eps', dpi=280)

log_gauss_error = [i for i in log_gauss_error if i is not None]
rnn_day_series_mape = [i for i in rnn_day_series_mape if i is not None]

# likelihood_hist
fig8 = plt.figure()
ax8 = fig8.add_subplot(1,1,1)
ax8.hist(log_gauss_error, bins=30, alpha = 0.5, label = 'rnn_likelihood')
# ax8.hist(s_arima_log_lkhd , bins=30, alpha = 0.5, label = 'sarima_likelihood')
ax8.legend()
ax8.vlines([-thrd], 0, 40, "r", linestyles='dashed')
ax8.set_title('lkhd per day histogram')
fig8.savefig(path_fig + 'seq2seq_sarima_lkhd_p_d_hist.pdf', dpi = 250)

# mape_hist
fig4 = plt.figure()
ax4 = fig4.add_subplot(1,1,1)
ax4.hist(np.array(rnn_day_series_mape, dtype = float).reshape(-1), bins=30, alpha = 0.5, label = 'rnn mape')
# ax4.hist(s_arima_day_series_mape, bins = 30, alpha = 0.5, label = 'sarima mape')
ax4.legend()
ax4.set_title('mape per day histogram')
fig4.savefig(path_fig + 'seq2seq_sarima_mape_p_d_hist.pdf', dpi = 250)

# print(s_arima_err_p_h)
print()
print(series_error)

# np_s_arima_err_p_h = np.reshape(s_arima_err_p_h, -1)
np_series_error = np.reshape(series_error, -1)

# print(np_s_arima_err_p_h)
print()
print(np_series_error)

# fig10 = plt.figure(figsize=(20, 8))
# ax10 = fig10.add_subplot(1,1,1)
# rnn = ax10.plot(range(len(np_series_error)), np_series_error, label='rnn_error')
# # sarima = ax10.plot(range(len(np_s_arima_err_p_h)), np_s_arima_err_p_h, label = 'sarima_error')
# ax10.legend()
# # ax10.hlines([-thrd], 0, len(s_arima_log_lkhd), "r", linestyles='dashed')
# ax10.set_xlabel('hour')
# ax10.set_ylabel('error')
# fig10.savefig(path_fig + 'seq2seq_sarima_error_series.eps', dpi=280)

# fig9 = plt.figure()
# ax9 = fig9.add_subplot(1,1,1)
# # ax9.hist(np_s_arima_err_p_h, bins=200, alpha = 0.5, label = 'rnn error')
# ax9.hist(np_series_error, bins=200, alpha = 0.5, label = 'sarima error')
# ax9.legend()
# ax9.set_title('error per hour histogram')
# fig9.savefig(path_fig + 'seq2seq_sarima_error_p_h_hist.pdf', dpi = 250)
